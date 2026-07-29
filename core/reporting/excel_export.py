"""Publication-ready Excel report builder.

Consumes only already-computed data (locked plan, analysis results, Table 1
DataFrame, STROBE report, KM plot).  Never computes or alters any statistic.
"""

from __future__ import annotations

import json
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

from core.reporting import format_label as _format_label


from core.database import get_connection, DATA_ROOT

# ── Style constants ────────────────────────────────────────────────────

NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E78")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E78")
BODY_FONT = Font(name="Calibri", size=11)
MONO_FONT = Font(name="Consolas", size=10)
ZEBRA_FILL = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
SUBTLE_HEADER_FILL = PatternFill(start_color="E8EBF0", end_color="E8EBF0", fill_type="solid")


def _style_header_row(ws, row: int, max_col: int):
    """Apply navy header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = WHITE_FONT
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_zebra(ws, start_row: int, end_row: int, max_col: int):
    """Apply alternating row shading."""
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = ZEBRA_FILL


def _autofit_columns(ws, min_width: int = 10, max_width: int = 55):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


def generate_excel_report(
    study_id: str,
    output_path: Optional[str | Path] = None,
) -> Path:
    """Generate a publication-ready 4-tab Excel workbook for a completed study.

    Parameters
    ----------
    study_id : str
    output_path : str or Path, optional
        Defaults to ``data/studies/{study_id}/study_report.xlsx``.

    Returns
    -------
    Path to the generated workbook.

    Raises
    ------
    RuntimeError
        If no analysis results exist or no locked plan is found.
    """
    conn = get_connection(study_id)

    # ── Study metadata ──────────────────────────────────────────────
    cur = conn.execute("SELECT * FROM studies WHERE id=?", (study_id,))
    study = cur.fetchone()
    if not study:
        conn.close()
        raise RuntimeError(f"Study '{study_id}' not found.")

    # ── Locked plan ─────────────────────────────────────────────────
    locked_paths = sorted(DATA_ROOT.glob(f"{study_id}/study_plan.v*.locked.json"))
    if not locked_paths:
        conn.close()
        raise RuntimeError(f"No locked plan found for study '{study_id}'.")
    latest_plan = json.loads(locked_paths[-1].read_text())

    # ── Analysis results ────────────────────────────────────────────
    cur = conn.execute(
        "SELECT * FROM analysis_results WHERE study_id=? ORDER BY id",
        (study_id,),
    )
    analyses = cur.fetchall()
    if not analyses:
        conn.close()
        raise RuntimeError(f"No analysis results found for study '{study_id}'.")

    conn.close()

    # ── STROBE report ───────────────────────────────────────────────
    from core.reporting.strobe_checklist import generate_report
    strobe_text = generate_report(study_id)
    strobe_satisfied = f"{strobe_text.count('✓')}/{strobe_text.count('Item')}"

    # ── Build workbook ──────────────────────────────────────────────
    wb = Workbook()
    ws_exec = wb.active
    temp_files: list[Path] = []

    _build_tab1_summary(ws_exec, study_id, study, latest_plan, analyses, strobe_satisfied, temp_files)
    _build_tab2_table1(wb, study_id)
    _build_tab3_analyses(wb, study_id, analyses)
    _build_tab4_audit(wb, study_id, latest_plan)

    # Auto-fit and polish
    for ws in wb.worksheets:
        ws.sheet_properties.showGridLines = False
        _autofit_columns(ws)

    if output_path is None:
        output_path = DATA_ROOT / study_id / "study_report.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(output_path))

    # Clean up temp files
    for tf in temp_files:
        tf.unlink(missing_ok=True)

    return output_path


# ── Tab 1: Executive Summary ────────────────────────────────────────────

def _build_tab1_summary(ws, study_id, study, plan_data, analyses, strobe_satisfied, temp_files: list[Path] = None):
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "1F4E78"

    # Title
    ws.cell(row=1, column=1, value="STUDY SUMMARY REPORT").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Metadata block
    meta = [
        ("Study ID", study_id),
        ("Study Title", study["name"] or "N/A"),
        ("Design Type", (study["study_type"] or "cohort").replace("_", " ")),
        ("Locked Plan", f"v{plan_data.get('version', '?')} (locked: {plan_data.get('locked_at', '?')[:10]})"),
        ("STROBE Compliance", f"{strobe_satisfied} items satisfied"),
        ("Created", (study["created_at"] or "")[:10]),
    ]
    for i, (label, value) in enumerate(meta):
        ws.cell(row=3 + i, column=1, value=label).font = HEADER_FONT
        ws.cell(row=3 + i, column=2, value=value).font = BODY_FONT

    # ── Primary analysis table ──────────────────────────────────────────
    r = 11
    ws.cell(row=r, column=1, value="PRIMARY OUTCOME").font = HEADER_FONT
    r += 1

    tbl_headers = ["Test Name", "Statistic", "P-Value", "Pre-Registered", "Result", "Rationale"]
    for ci, h in enumerate(tbl_headers, 1):
        ws.cell(row=r, column=ci, value=h)
    _style_header_row(ws, r, len(tbl_headers))
    r += 1

    pre_reg = [a for a in analyses if a["is_pre_registered"]]
    post_hoc = [a for a in analyses if not a["is_pre_registered"]]

    # Pre-registered rows
    pre_start = r
    for a in analyses:
        if not a["is_pre_registered"]:
            continue
        test_name = a["test_name"]
        stat = f"{a['statistic']:.4f}" if a["statistic"] is not None else "N/A"
        pv = f"{a['p_value']:.4f}" if a["p_value"] is not None else "N/A"
        pre = "YES"
        status_data = json.loads(a["status_json"]) if a["status_json"] else {}
        status = status_data.get("status", "completed")
        from core.planning.diagnostics import check_violation
        has_viol, viol_summary, _ = check_violation(dict(a))
        sig = "Significant" if (a["p_value"] is not None and a["p_value"] < 0.05) else "Not Significant"
        if has_viol:
            sig = f"Significant — ⚠ assumption violation"
        vals = [test_name, stat, pv, pre, sig, "Pre-registered primary protocol"]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=r, column=ci, value=v).font = BODY_FONT
            ws.cell(row=r, column=ci).border = THIN_BORDER
        r += 1

    _apply_zebra(ws, pre_start, r - 1, len(tbl_headers))

    # Post-hoc section
    if post_hoc:
        r += 1
        ws.cell(row=r, column=1, value="EXPLORATORY / POST-HOC ANALYSES").font = HEADER_FONT
        r += 1

        ph_headers = ["Test Name", "Statistic", "P-Value", "Result", "Rationale", "Reason"]
        for ci, h in enumerate(ph_headers, 1):
            ws.cell(row=r, column=ci, value=h)
        _style_header_row(ws, r, len(ph_headers))
        r += 1

        ph_start = r
        for a in post_hoc:
            test_name = a["test_name"]
            stat = f"{a['statistic']:.4f}" if a["statistic"] is not None else "N/A"
            pv = f"{a['p_value']:.4f}" if a["p_value"] is not None else "N/A"
            sig = "Significant" if (a["p_value"] is not None and a["p_value"] < 0.05) else "Not Significant"
            prov = {}
            if a["provenance_json"]:
                try:
                    prov = json.loads(a["provenance_json"])
                except (TypeError, json.JSONDecodeError):
                    pass
            rationale = prov.get("rationale", "")
            reason = prov.get("amendment_reason", "")
            label = f"{test_name} — {rationale}" if rationale else test_name
            vals = [label, stat, pv, sig, rationale, reason]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=r, column=ci, value=v).font = BODY_FONT
                ws.cell(row=r, column=ci).border = THIN_BORDER
            r += 1

        _apply_zebra(ws, ph_start, r - 1, len(ph_headers))

    # Embed KM plot as PNG
    r += 2
    try:
        from core.reporting.plots import generate_km_plot
        km_result = next(
            (a for a in analyses
             if a["test_name"] == "kaplan_meier_logrank"
             and json.loads(a["status_json"] or "{}").get("status") == "completed"),
            None,
        )
        if km_result:
            test_id = km_result["id"]
            tmp_png = Path(tempfile.mkstemp(suffix=".png")[1])
            generate_km_plot(study_id, test_id=test_id, output_path=tmp_png, fmt="png", style="clean")
            if temp_files is not None:
                temp_files.append(tmp_png)
            img = XlImage(str(tmp_png))
            img.width, img.height = 600, 380
            ws.add_image(img, f"A{r}")
            r += 25
    except Exception:
        pass


# ── Tab 2: Table 1 — Baseline Characteristics ───────────────────────────

VALUE_MAPPINGS = {
    "sex": {"f": "Female", "m": "Male"},
    "high_risk_cytogenetics": {"yes": "Yes", "no": "No"},
    "iss_stage": {"I": "Stage I", "II": "Stage II", "III": "Stage III"},
    "treatment_arm": {"A": "Test Arm (A)", "B": "Placebo Arm (B)"},
}


def _build_tab2_table1(wb, study_id):
    ws = wb.create_sheet(title="Table 1 - Baseline")
    ws.sheet_properties.tabColor = "1F4E78"

    from core.stats.descriptive import generate_table1
    from core.database import get_connection

    groupby = "treatment_arm"
    tbl = generate_table1(study_id, groupby=groupby)

    if tbl.empty:
        ws.cell(row=1, column=1, value="Table 1 not yet computed.").font = BODY_FONT
        return

    # Flatten columns — drop the "Missing" column
    cols = list(tbl.columns)
    if hasattr(tbl.columns, "levels") or (cols and isinstance(cols[0], tuple)):
        cols = [str(c[-1]).strip() if isinstance(c, tuple) else str(c) for c in cols]
    if cols and cols[0].lower() in ("missing", ""):
        cols = cols[1:]
        tbl = tbl.iloc[:, 1:]

    # Get sample sizes for header labels
    conn = get_connection(study_id)
    raw = f"raw_{study_id}"
    total_n = conn.execute(f"SELECT COUNT(*) as n FROM {raw}").fetchone()["n"]
    arm_counts = {}
    if groupby:
        col_names = [r["name"] for r in conn.execute(f"PRAGMA table_info({raw})")]
        if groupby in col_names:
            cur = conn.execute(f'SELECT "{groupby}" as arm, COUNT(*) as n FROM {raw} GROUP BY arm')
            arm_counts = {r["arm"]: r["n"] for r in cur.fetchall()}
    conn.close()

    # Build header row with N counts
    header_parts = ["Characteristic", f"Overall (N={total_n})"]
    for c in cols[1:]:
        n = arm_counts.get(c, "")
        header_parts.append(f"{c} (N={n})" if n else c)
    header_row = header_parts
    ncols = len(header_row)

    for ci, h in enumerate(header_row, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, ncols)

    # Process the original MultiIndex
    orig_idx = list(tbl.index) if hasattr(tbl.index, "levels") else []
    if not orig_idx:
        return

    data_row = 2
    prev_var_raw = ""
    for ri in range(len(tbl)):
        entry = orig_idx[ri]
        raw_label = str(entry[0]).strip() if isinstance(entry, tuple) else str(entry).strip()
        cat_level = str(entry[1]).strip() if isinstance(entry, tuple) and entry[1] else ""

        if raw_label.lower() == "n":
            continue

        is_continuous = not cat_level
        var_base = raw_label.split(",")[0].strip() if "," in raw_label else raw_label

        if var_base.lower().replace(" ", "_") == "treatment_arm":
            continue

        if not is_continuous:
            var_key = var_base.lower().replace(" ", "_").replace("-", "_")
            if var_key != prev_var_raw:
                var_parts = raw_label.split(",", 1)
                var_name = _format_label(var_parts[0]).strip()
                suffix = "," + var_parts[1] if len(var_parts) > 1 else ""
                group_name = var_name + suffix
                cell = ws.cell(row=data_row, column=1, value=group_name)
                cell.font = Font(name="Calibri", bold=True, size=11)
                data_row += 1
                prev_var_raw = var_key

            mapping = VALUE_MAPPINGS.get(var_key, {})
            clean_cat = mapping.get(cat_level, mapping.get(cat_level.lower(), cat_level))
            cell = ws.cell(row=data_row, column=1, value=f"  {clean_cat}")
            cell.font = BODY_FONT
        else:
            prev_var_raw = ""
            var_parts = raw_label.split(",", 1)
            var_name = _format_label(var_parts[0]).strip()
            suffix = "," + var_parts[1] if len(var_parts) > 1 else ""
            clean_label = var_name + suffix
            cell = ws.cell(row=data_row, column=1, value=clean_label)
            cell.font = BODY_FONT

        for ci in range(ncols - 1):
            if ci < len(tbl.columns):
                raw_val = tbl.iloc[ri, ci]
                val_str = str(raw_val) if not pd.isna(raw_val) else ""
                ws.cell(row=data_row, column=ci + 2, value=val_str).font = BODY_FONT

        for ci in range(1, ncols + 1):
            ws.cell(row=data_row, column=ci).border = THIN_BORDER

        data_row += 1

    _apply_zebra(ws, 2, data_row - 1, ncols)


# ── Tab 3: Statistical Analyses ─────────────────────────────────────────

def _build_tab3_analyses(wb, study_id, analyses):
    import json
    ws = wb.create_sheet(title="Statistical Analyses")
    ws.sheet_properties.tabColor = "1F4E78"

    locked_paths = sorted(DATA_ROOT.glob(f"{study_id}/study_plan.v*.locked.json"))
    plan_data = None
    if locked_paths:
        plan_data = json.loads(locked_paths[-1].read_text())

    primary_comparison = ""
    var_map = {}
    if plan_data:
        primary_comparison = plan_data.get("primary_comparison", "")
        for t in plan_data.get("planned_tests", []):
            var_map[t.get("test_name", "")] = t.get("variable_name", "")
        for t in plan_data.get("post_hoc_tests", []):
            var_map[t.get("test_name", "")] = t.get("variable_name", "")

    # Pre-fetch covariate results
    conn = get_connection(study_id)
    covariate_map: dict[int, list[dict]] = {}
    result_ids = [a["id"] for a in analyses]
    if result_ids:
        placeholders = ",".join("?" for _ in result_ids)
        cur = conn.execute(
            f"SELECT * FROM analysis_covariate_results WHERE result_id IN ({placeholders}) ORDER BY id",
            result_ids,
        )
        for row in cur.fetchall():
            covariate_map.setdefault(row["result_id"], []).append({
                "covariate": row["covariate"],
                "hr": row["hr"],
                "ci_lower": row["ci_lower"],
                "ci_upper": row["ci_upper"],
                "wald_p": row["wald_p"],
            })
    conn.close()

    headers = [
        "Analysis ID", "Test Name", "Comparison", "Variable",
        "N Analyzed", "Statistic", "P-Value", "Adj. P-Value",
        "95% CI (Lower)", "95% CI (Upper)",
        "LR Test P", "Concordance",
        "Status", "Pre-Registered",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))

    r = 2
    footnote_idx = 0
    footnotes: list[str] = []
    for a in analyses:
        status_data = json.loads(a["status_json"]) if a["status_json"] else {}
        status = status_data.get("status", "completed")
        from core.planning.diagnostics import check_violation
        has_viol, viol_summary, _ = check_violation(dict(a))
        display_status = status
        if has_viol:
            footnote_idx += 1
            marker = chr(0x2460 + footnote_idx - 1)  # circled digits: ① ② ...
            display_status = f"{status} {marker}"
            footnotes.append(f"{marker} {viol_summary}")
        sc = json.loads(a["sample_counts_json"]) if a["sample_counts_json"] else {}
        n = sc.get("n_analyzed", "")
        test_name = a["test_name"]
        comparison = primary_comparison
        variable = var_map.get(test_name, "")
        if variable and variable != "pfs_days" and variable != "os_days":
            v_clean = _format_label(variable)
            comparison = f"{v_clean} by Treatment Arm"
        ci_lower_str = f"{a['ci_lower']:.4f}" if a["ci_lower"] is not None else "N/A"
        ci_upper_str = f"{a['ci_upper']:.4f}" if a["ci_upper"] is not None else "N/A"
        lr_p_str = ""
        c_index_str = ""
        try:
            if a["lr_test_p"] is not None:
                lr_p_str = f"{a['lr_test_p']:.4f}"
        except (KeyError, IndexError, TypeError):
            pass
        try:
            if a["concordance_index"] is not None:
                c_index_str = f"{a['concordance_index']:.3f}"
        except (KeyError, IndexError, TypeError):
            pass
        vals = [
            a["id"],
            test_name,
            comparison,
            variable,
            n,
            f"{a['statistic']:.4f}" if a["statistic"] is not None else "",
            f"{a['p_value']:.4f}" if a["p_value"] is not None else "",
            f"{a['adjusted_p_value']:.4f}" if a["adjusted_p_value"] is not None else "",
            ci_lower_str,
            ci_upper_str,
            lr_p_str,
            c_index_str,
            display_status,
            "YES" if a["is_pre_registered"] else "NO  [POST-HOC]",
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=r, column=ci, value=v).font = BODY_FONT
            ws.cell(row=r, column=ci).border = THIN_BORDER
        r += 1

        # Per-covariate sub-rows for Cox PH
        cov_rows = covariate_map.get(a["id"], [])
        if cov_rows:
            cov_headers = ["", "Covariate", "HR", "CI Lower", "CI Upper", "P-Value"]
            cov_start = r
            for ci, h in enumerate(cov_headers, 1):
                ws.cell(row=r, column=ci, value=h).font = Font(name="Calibri", italic=True, size=10, color="555555")
            r += 1
            for cr in cov_rows:
                ws.cell(row=r, column=2, value=cr["covariate"]).font = BODY_FONT
                ws.cell(row=r, column=3, value=f"{cr['hr']:.4f}" if cr.get("hr") is not None else "").font = BODY_FONT
                ws.cell(row=r, column=4, value=f"{cr['ci_lower']:.4f}" if cr.get("ci_lower") is not None else "").font = BODY_FONT
                ws.cell(row=r, column=5, value=f"{cr['ci_upper']:.4f}" if cr.get("ci_upper") is not None else "").font = BODY_FONT
                ws.cell(row=r, column=6, value=f"{cr['wald_p']:.4f}" if cr.get("wald_p") is not None else "").font = BODY_FONT
                r += 1

    _apply_zebra(ws, 2, r - 1, len(headers))

    # Footnotes for assumption violations
    for note in footnotes:
        ws.cell(row=r, column=1, value=note).font = Font(name="Calibri", italic=True, size=9, color="C0392B")
        r += 1


# ── Tab 4: Audit & Hash Manifest ────────────────────────────────────────

def _build_tab4_audit(wb, study_id, plan_data):
    ws = wb.create_sheet(title="Audit & Hash Manifest")
    ws.sheet_properties.tabColor = "1F4E78"

    ws.cell(row=1, column=1, value="FILE").font = WHITE_FONT
    ws.cell(row=1, column=1).fill = NAVY_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.cell(row=1, column=2, value="SHA-256 HASH").font = WHITE_FONT
    ws.cell(row=1, column=2).fill = NAVY_FILL
    ws.cell(row=1, column=2).border = THIN_BORDER

    from core.provenance.hashing import sha256 as _sha256, canonical_json as _canonical_json, compute_raw_data_hash

    conn = get_connection(study_id)
    raw_data_hash = compute_raw_data_hash(study_id)

    locked_plan_hash = plan_data.get("content_hash", "N/A")

    cur = conn.execute(
        "SELECT * FROM analysis_results WHERE study_id=? ORDER BY id", (study_id,)
    )
    results = []
    for r in cur.fetchall():
        row = dict(r)
        for jf in ("variable_ids_used", "effect_size_json",
                   "sample_counts_json", "status_json", "provenance_json",
                   "ph_diagnostics_json"):
            if row.get(jf):
                try:
                    row[jf] = json.loads(row[jf])
                except (TypeError, json.JSONDecodeError):
                    pass
        results.append(row)
    conn.close()
    results_json = _canonical_json(results)
    results_hash = _sha256(results_json)

    COMPOSITE_SEP = "||"
    composite_hash = _sha256(COMPOSITE_SEP.join([raw_data_hash, locked_plan_hash, results_hash]))

    rows = [
        ("Study Plan (locked)", locked_plan_hash),
        ("Raw Data (ingested)", raw_data_hash),
        ("Analysis Results", results_hash),
        ("Composite Bundle Hash", composite_hash),
    ]

    for ri, (label, hash_val) in enumerate(rows):
        r = ri + 2
        ws.cell(row=r, column=1, value=label).font = MONO_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=hash_val).font = MONO_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER


import pandas as pd
