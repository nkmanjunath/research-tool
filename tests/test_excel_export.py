"""Tests for Excel report export."""

from __future__ import annotations

import json
import shutil
import tempfile
from pathlib import Path

import pytest

from core.database import get_connection, init_db, DATA_ROOT
from core.planning.lock import lock_plan
from core.planning.study_plan import StudyPlan
from core.reporting.excel_export import generate_excel_report


STUDY_ID = "test_excel"


@pytest.fixture(autouse=True)
def _setup():
    p = DATA_ROOT / STUDY_ID
    if p.exists():
        shutil.rmtree(p)
    p.mkdir(parents=True)

    conn = get_connection(STUDY_ID)
    init_db(conn)
    conn.execute(
        "INSERT OR REPLACE INTO studies (id, name, created_at, data_dir, study_type, is_locked) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (STUDY_ID, "Excel Test", "2025-01-01", str(DATA_ROOT / STUDY_ID), "cohort", 0),
    )
    raw = f"raw_{STUDY_ID}"
    conn.execute(f"CREATE TABLE IF NOT EXISTS {raw} (row_id INTEGER PRIMARY KEY, age TEXT, response TEXT, treatment_arm TEXT)")
    conn.execute(f"INSERT INTO {raw} (age, response, treatment_arm) VALUES ('65', 'CR', 'A')")
    conn.execute(f"INSERT INTO {raw} (age, response, treatment_arm) VALUES ('70', 'PR', 'B')")
    conn.execute("DELETE FROM variables WHERE study_id=?", (STUDY_ID,))
    conn.execute("INSERT INTO variables (study_id, column_name, role, data_type) VALUES (?, 'age', 'baseline', 'continuous')", (STUDY_ID,))
    conn.execute("INSERT INTO variables (study_id, column_name, role, data_type) VALUES (?, 'response', 'outcome', 'categorical')", (STUDY_ID,))
    conn.execute("INSERT INTO variables (study_id, column_name, role, data_type) VALUES (?, 'treatment_arm', 'baseline', 'categorical')", (STUDY_ID,))
    conn.commit()
    conn.close()

    from core.masking.gate import seal_outcomes
    seal_outcomes(STUDY_ID)
    yield
    if p.exists():
        shutil.rmtree(p)


def _setup_plan_and_result():
    plan = StudyPlan(study_id=STUDY_ID, study_type="cohort", primary_comparison="test",
                     planned_tests=[{"variable_name": "response", "test_name": "chi_square"}])
    lock_plan(STUDY_ID, plan)
    conn = get_connection(STUDY_ID)
    init_db(conn)
    conn.execute(
        "INSERT INTO analysis_results (id, study_id, test_name, statistic, p_value, status_json, "
        "sample_counts_json, is_pre_registered, computed_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)",
        (1, STUDY_ID, "chi_square", 5.2, 0.022,
         json.dumps({"status": "completed"}),
         json.dumps({"n_total": 2, "n_analyzed": 2, "n_excluded": 0}),
         "2025-01-01T00:00:00"),
    )
    conn.commit()
    conn.close()


def test_excel_report_generates_valid_xlsx():
    """A completed study should produce a valid .xlsx file with 4 sheets."""
    _setup_plan_and_result()
    out = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    try:
        result = generate_excel_report(STUDY_ID, output_path=out)
        assert result.exists()
        assert result.suffix == ".xlsx"
        import openpyxl
        wb = openpyxl.load_workbook(str(result))
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 4, f"Expected 4 sheets, got {len(sheet_names)}: {sheet_names}"
        # Check expected tab names
        assert "Executive Summary" in sheet_names
        assert "Table 1 - Baseline" in sheet_names
        assert "Statistical Analyses" in sheet_names
        assert "Audit & Hash Manifest" in sheet_names
    finally:
        out.unlink()


def test_excel_report_rejects_no_results():
    """Must fail with a clear error if no analysis results exist."""
    # Create a locked plan first so it passes the plan check
    from core.masking.gate import seal_outcomes
    seal_outcomes(STUDY_ID)
    plan = StudyPlan(study_id=STUDY_ID, study_type="cohort", primary_comparison="test",
                     planned_tests=[{"variable_name": "response", "test_name": "chi_square"}])
    lock_plan(STUDY_ID, plan)
    out = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    with pytest.raises(RuntimeError, match="No analysis results"):
        generate_excel_report(STUDY_ID, output_path=out)
    out.unlink()


def test_excel_cox_ph_covariate_alignment():
    """B13 regression: Cox PH covariate rows must align under Col 4 (Variable), Col 6 (HR), Col 7 (P-Value), Col 9/10 (CI)."""
    plan = StudyPlan(
        study_id=STUDY_ID,
        study_type="cohort",
        primary_comparison="test comparison",
        planned_tests=[{"variable_name": "age", "test_name": "cox_ph_model"}],
    )
    lock_plan(STUDY_ID, plan)
    conn = get_connection(STUDY_ID)
    init_db(conn)
    conn.execute(
        "INSERT INTO analysis_results (id, study_id, test_name, statistic, p_value, status_json, "
        "sample_counts_json, is_pre_registered, computed_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)",
        (1, STUDY_ID, "cox_ph_model", 3.45, 0.15,
         json.dumps({"status": "completed"}),
         json.dumps({"n_total": 2, "n_analyzed": 2, "n_excluded": 0}),
         "2025-01-01T00:00:00"),
    )
    conn.execute(
        "INSERT INTO analysis_covariate_results (result_id, covariate, hr, ci_lower, ci_upper, wald_p) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (1, "age", 1.05, 0.98, 1.12, 0.1234),
    )
    conn.commit()
    conn.close()

    out = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    try:
        generate_excel_report(STUDY_ID, output_path=out)
        import openpyxl
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Statistical Analyses"]

        # Row 1 is parent header
        assert ws.cell(1, 4).value == "Variable"
        assert ws.cell(1, 6).value == "Statistic"
        assert ws.cell(1, 7).value == "P-Value"
        assert ws.cell(1, 9).value == "95% CI (Lower)"
        assert ws.cell(1, 10).value == "95% CI (Upper)"

        # Row 3 is sub-header for Cox PH covariates
        assert ws.cell(3, 4).value == "Covariate"
        assert ws.cell(3, 6).value == "HR"
        assert ws.cell(3, 7).value == "P-Value"
        assert ws.cell(3, 9).value == "CI Lower"
        assert ws.cell(3, 10).value == "CI Upper"

        # Row 4 is the covariate data row
        assert ws.cell(4, 4).value == "age"
        assert ws.cell(4, 6).value == "1.0500"
        assert ws.cell(4, 7).value == "0.1234"
        assert ws.cell(4, 9).value == "0.9800"
        assert ws.cell(4, 10).value == "1.1200"
    finally:
        out.unlink()

